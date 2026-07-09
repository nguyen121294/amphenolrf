import openpyxl
import os
import sys

def analyze():
    s_path = r"C:\Users\jackn\spyder\amphenolrf\docs\Stest.xlsx"
    r_path = r"C:\Users\jackn\spyder\amphenolrf\docs\Rtest.xlsx"

    if not os.path.exists(s_path):
        print(f"File not found: {s_path}")
        return
    if not os.path.exists(r_path):
        print(f"File not found: {r_path}")
        return

    # Load resources
    r_wb = openpyxl.load_workbook(r_path, data_only=True)
    r_sheet = r_wb.worksheets[0]
    line_capacities = {}
    default_capacity = [10.5, 10.5, 10.5, 10.5, 10.5, 10.5, 10.5]
    
    # Parse line capacities from Rtest.xlsx
    for row in r_sheet.iter_rows(values_only=True):
        if not row or len(row) == 0:
            continue
        cell_0 = str(row[0] or "").strip().upper()
        if cell_0.startswith("L") and cell_0[1:].isdigit():
            capacities = []
            for c in range(1, 8):
                val = 0.0
                if c < len(row) and row[c] is not None:
                    try:
                        val = float(row[c])
                    except:
                        pass
                capacities.append(val)
            line_capacities[cell_0] = capacities

    # Fill default active lines
    active_lines = ["L1", "L2", "L3", "L4", "L5", "L6", "L7", "L8", "L9", "L10", "L11", "L12", "L14", "L15", "L16"]
    for line in active_lines:
        if line not in line_capacities:
            if line == "L4":
                line_capacities[line] = [10.5, 10.5, 8.0, 10.5, 10.5, 10.5, 10.5]
            elif line == "L5":
                line_capacities[line] = [10.5, 10.5, 10.5, 10.5, 0.0, 10.5, 10.5]
            else:
                line_capacities[line] = default_capacity

    # Load shipping list (Stest.xlsx)
    s_wb = openpyxl.load_workbook(s_path, data_only=True)
    found_sheet = None
    for name in s_wb.sheetnames:
        if name.lower() == "shipping list":
            found_sheet = s_wb[name]
            break
    if not found_sheet:
        found_sheet = s_wb.worksheets[0]
    
    sheet = found_sheet
    print(f"Analyzing sheet: {sheet.title}")

    # Parse rows starting from dataRow=3 (index 2 in 0-based)
    rows = list(sheet.iter_rows(values_only=True))
    data_rows = rows[2:] # 0-indexed index 2 is row 3
    
    parsed_items = []
    
    def text_val(val):
        if val is None:
            return ""
        return str(val).strip()

    def num_val(val, fallback=0.0):
        if val is None:
            return fallback
        try:
            return float(val)
        except:
            return fallback

    def normalize(val):
        text = text_val(val).lower()
        # Simple normalization
        import unicodedata
        text = unicodedata.normalize('NFD', text)
        text = "".join([c for c in text if unicodedata.category(c) != 'Mn'])
        text = text.replace('đ', 'd')
        return text

    def is_not_no(val):
        return val != "no"

    for r_idx, row in enumerate(data_rows):
        # excel row number is r_idx + 3
        row_number = r_idx + 3
        if not row or len(row) < 10:
            continue
        
        # Column mappings in code:
        # C is item (index 2)
        # B is desc (index 1)
        # I is mo (index 8)
        # J is quantity (index 9)
        # T is uph (index 19)
        # H is status (index 7)
        # G is pmc (index 6)
        # BE is solderPN (index 56)
        # BG is inkPrinting (index 58)
        # BH is letterRolling (index 59)
        # BI is tapeReel (index 60)
        # AD is promiseDate (index 29)
        # Z is shipDate (index 25)
        
        item = text_val(row[2])
        desc = text_val(row[1])
        mo = text_val(row[8])
        quantity = num_val(row[9])
        uph = num_val(row[19])
        status = text_val(row[7])
        pmc = text_val(row[6])
        
        # Check if row is empty
        if not quantity and not mo and not item and not desc:
            continue
        
        # Apply user UPH override
        # User said: "Tôi cho hiệu xuất UPH bằng 1, các UPH bằng 0 tôi đều áp bằng 300"
        if uph <= 0:
            uph = 300.0

        solder_pn = text_val(row[56]) if len(row) > 56 else ""
        ink_printing = text_val(row[58]) if len(row) > 58 else ""
        letter_rolling = text_val(row[59]) if len(row) > 59 else ""
        tape_reel = text_val(row[60]) if len(row) > 60 else ""
        
        norm_desc = normalize(desc)
        norm_status = normalize(status)
        norm_pmc = normalize(pmc)
        norm_solder = normalize(solder_pn)
        norm_ink = normalize(ink_printing)
        norm_letter = normalize(letter_rolling)
        norm_tape = normalize(tape_reel)
        
        pmc_score = 6 if "uu tien 1" in norm_pmc else 1
        sap_score = 5 if "sap" in norm_desc or is_not_no(norm_letter) else 1
        kap_score = 5 if "kap" in norm_desc or is_not_no(norm_ink) else 1
        
        # fixLatheScore=True startsWith("34-")
        is_lathe = item.upper().startswith("34-")
        lathe_score = 5 if is_lathe else 1
        
        reel_score = 4 if "smp" in norm_desc or "stpk" in norm_desc or is_not_no(norm_tape) else 1
        solder_score = 3 if norm_solder == "yes" else 1
        large_qty_score = 2 if quantity > 5000 else 1
        
        material_rank = 1
        try:
            h_val = int(status.strip())
            if h_val == 3: material_rank = 4
            elif h_val == 2: material_rank = 3
            elif h_val == 1: material_rank = 2
            elif h_val == 0: material_rank = 1
        except:
            if "ready material" in norm_status:
                material_rank = 4
            elif "co lieu" in norm_status:
                material_rank = 3
                
        priority_score = pmc_score * sap_score * kap_score * lathe_score * reel_score * solder_score * large_qty_score
        
        # Date sorting
        # Let's just use string or simple numeric conversion for dates
        promise_date = row[29] if len(row) > 29 else ""
        ship_date = row[25] if len(row) > 25 else ""
        
        def date_val(val):
            # Simplistic parser
            if isinstance(val, (int, float)):
                return float(val)
            return 999999.0
            
        parsed_items.append({
            'row_num': row_number,
            'mo': mo,
            'item': item,
            'desc': desc,
            'quantity': quantity,
            'uph': uph,
            'priority_score': priority_score,
            'material_rank': material_rank,
            'pmc_score': pmc_score,
            'promise_date': date_val(promise_date),
            'ship_date': date_val(ship_date),
            'original_order': len(parsed_items)
        })

    # Filter schedulable (quantity > 0 and uph > 0)
    schedulable = [it for it in parsed_items if it['quantity'] > 0 and it['uph'] > 0]
    blocked = [it for it in parsed_items if it['quantity'] > 0 and it['uph'] <= 0]
    
    print(f"Total parsed items with quantity > 0: {len(schedulable) + len(blocked)}")
    print(f"Schedulable items: {len(schedulable)}")
    print(f"Blocked items (UPH <= 0): {len(blocked)}")

    # Group by item code
    groups_map = {}
    for item_data in schedulable:
        group_key = item_data['item']
        if group_key not in groups_map:
            groups_map[group_key] = {
                'key': group_key,
                'items': [],
                'best_item': item_data,
                'first_order': item_data['original_order']
            }
        groups_map[group_key]['items'].append(item_data)
        groups_map[group_key]['first_order'] = min(groups_map[group_key]['first_order'], item_data['original_order'])
        
        # Compare items helper
        def is_better(a, b):
            # Compare a and b, return True if a is better than b
            if a['priority_score'] != b['priority_score']:
                return a['priority_score'] > b['priority_score']
            if a['material_rank'] != b['material_rank']:
                return a['material_rank'] > b['material_rank']
            if a['pmc_score'] != b['pmc_score']:
                return a['pmc_score'] > b['pmc_score']
            if a['promise_date'] != b['promise_date']:
                return a['promise_date'] < b['promise_date']
            if a['ship_date'] != b['ship_date']:
                return a['ship_date'] < b['ship_date']
            return a['original_order'] < b['original_order']
            
        if is_better(item_data, groups_map[group_key]['best_item']):
            groups_map[group_key]['best_item'] = item_data

    plan_groups = list(groups_map.values())
    
    # Sort plan groups
    def group_key_fn(g):
        best = g['best_item']
        # Return sorting tuple
        # JS sort: b.priorityScore - a.priorityScore (desc)
        # So we negate values for descending sort in python
        return (
            -best['priority_score'],
            -best['material_rank'],
            -best['pmc_score'],
            best['promise_date'],
            best['ship_date'],
            best['original_order']
        )
        
    plan_groups.sort(key=group_key_fn)

    print(f"Number of product groups (unique Items): {len(plan_groups)}")

    # Allocated times track
    allocated_times = {line: [0.0]*7 for line in active_lines}
    
    # Round-Robin assignment
    line_pointer = 0
    group_assignments = {}
    
    for g in plan_groups:
        assigned_line = active_lines[line_pointer]
        group_assignments[g['key']] = assigned_line
        line_pointer = (line_pointer + 1) % len(active_lines)
        
        # Find preferredDay
        capacities = line_capacities[assigned_line]
        preferred_day = 0
        for d in range(7):
            if capacities[d] - allocated_times[assigned_line][d] > 0.01:
                preferred_day = d
                break
                
        # Sort items inside group
        g['items'].sort(key=lambda x: (
            -x['priority_score'],
            -x['material_rank'],
            -x['pmc_score'],
            x['promise_date'],
            x['ship_date'],
            x['original_order']
        ))
        
        # Allocate each item
        days_to_allocate = list(range(preferred_day, 7)) # Our updated logic (all days from preferred_day to index 6)
        
        qty_efficiency = 1.0 # User stated: "Tôi cho hiệu xuất UPH bằng 1"
        lot_step = 500.0
        
        for item in g['items']:
            remaining = item['quantity']
            uph_eff = item['uph'] * qty_efficiency
            
            allocated_qty_days = [0.0]*7
            
            for day_idx in days_to_allocate:
                if remaining <= 0:
                    break
                cap = capacities[day_idx]
                used = allocated_times[assigned_line][day_idx]
                avail = max(cap - used, 0.0)
                
                if avail <= 0.001:
                    continue
                    
                max_qty = avail * uph_eff
                chunk = min(remaining, max_qty)
                
                # Lot step rounding
                if lot_step > 1 and remaining > lot_step and chunk >= lot_step:
                    rounded = max(lot_step, round(chunk / lot_step) * lot_step)
                    chunk = min(rounded, remaining, max_qty)
                else:
                    chunk = min(chunk, remaining)
                    
                if chunk > 0:
                    time_used = chunk / uph_eff
                    allocated_qty_days[day_idx] += chunk
                    allocated_times[assigned_line][day_idx] += time_used
                    remaining -= chunk
                    
            item['allocated_qty'] = sum(allocated_qty_days)
            item['missing'] = remaining
            item['assigned_line'] = assigned_line

    # Let's print summary stats for each Line
    print("\n--- LINE CAPACITY UTILIZATION SUMMARY ---")
    line_groups_count = {line: 0 for line in active_lines}
    line_total_qty_demanded = {line: 0.0 for line in active_lines}
    line_total_qty_scheduled = {line: 0.0 for line in active_lines}
    
    for g in plan_groups:
        line = group_assignments[g['key']]
        line_groups_count[line] += 1
        for item in g['items']:
            line_total_qty_demanded[line] += item['quantity']
            line_total_qty_scheduled[line] += item['allocated_qty']

    for line in active_lines:
        cap_sum = sum(line_capacities[line])
        used_sum = sum(allocated_times[line])
        demanded = line_total_qty_demanded[line]
        scheduled = line_total_qty_scheduled[line]
        missing = demanded - scheduled
        print(f"Line {line:3s}: Groups={line_groups_count[line]:2d} | Cap={cap_sum:5.1f}h | Used={used_sum:5.1f}h | Demanded={demanded:8.1f} | Scheduled={scheduled:8.1f} | Missing={missing:8.1f}")
        
    print("\n--- DETAILED MISSING ITEMS ON BUSY LINES ---")
    for g in plan_groups:
        line = group_assignments[g['key']]
        for item in g['items']:
            if item['missing'] > 0.01:
                print(f"Row {item['row_num']:3d} | Item: {item['item']:12s} | MO: {item['mo']:10s} | Qty: {item['quantity']:8.1f} | Missing: {item['missing']:8.1f} | Line: {item['assigned_line']}")
                print(f"         Line {item['assigned_line']} Hours Used: {[round(h, 2) for h in allocated_times[item['assigned_line']]]} out of Capacity: {line_capacities[item['assigned_line']]}")

if __name__ == "__main__":
    analyze()
