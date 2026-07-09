import openpyxl

def check_status_cols():
    s_path = r"C:\Users\jackn\spyder\amphenolrf\docs\Stest.xlsx"
    wb = openpyxl.load_workbook(s_path, data_only=True)
    sheet = wb.worksheets[0]
    
    print("Checking Status column values:")
    # Row 3 to 20
    for r in range(3, 21):
        h_val = sheet.cell(row=r, column=8).value # H (Col 8)
        ak_val = sheet.cell(row=r, column=37).value # AK (Col 37)
        an_val = sheet.cell(row=r, column=40).value # AN (Col 40)
        print(f"Row {r} | H: {h_val} | AK: {ak_val} | AN: {an_val}")

if __name__ == "__main__":
    check_status_cols()
