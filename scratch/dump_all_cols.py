import openpyxl

def dump_all_cols():
    s_path = r"C:\Users\jackn\spyder\amphenolrf\docs\Stest.xlsx"
    wb = openpyxl.load_workbook(s_path, data_only=True)
    sheet = wb.worksheets[0]
    
    # Row 2 (headers)
    row_2 = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    for idx, cell in enumerate(row_2):
        col_letter = openpyxl.utils.get_column_letter(idx + 1)
        print(f"Index {idx} ({col_letter}): {cell}")

if __name__ == "__main__":
    dump_all_cols()
