import openpyxl

def inspect_102629():
    s_path = r"C:\Users\jackn\spyder\amphenolrf\docs\Stest.xlsx"
    r_path = r"C:\Users\jackn\spyder\amphenolrf\docs\Rtest.xlsx"
    
    wb_s = openpyxl.load_workbook(s_path, data_only=True)
    sheet_s = wb_s.worksheets[0]
    
    print("--- Searching in Stest.xlsx ---")
    for r in range(3, sheet_s.max_row+1):
        item = str(sheet_s.cell(row=r, column=2).value or "").strip()
        mo = str(sheet_s.cell(row=r, column=9).value or "").strip()
        qty = sheet_s.cell(row=r, column=10).value
        uph = sheet_s.cell(row=r, column=20).value
        if "102629" in item:
            print(f"Row {r} | Item: {item} | MO: {mo} | Qty: {qty} | UPH: {uph}")
            
    wb_r = openpyxl.load_workbook(r_path, data_only=True)
    sheet_r = wb_r.worksheets[0]
    print("\n--- Searching in Rtest.xlsx ---")
    for r in range(2, sheet_r.max_row+1):
        item = str(sheet_r.cell(row=r, column=1).value or "").strip()
        line = str(sheet_r.cell(row=r, column=2).value or "").strip()
        uph = sheet_r.cell(row=r, column=3).value
        if "102629" in item:
            print(f"Row {r} | Item: {item} | Line: {line} | UPH: {uph}")

if __name__ == "__main__":
    inspect_102629()
