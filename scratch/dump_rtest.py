import openpyxl

def dump_rtest():
    path = r"C:\Users\jackn\spyder\amphenolrf\docs\Rtest.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    for sheet_name in wb.sheetnames:
        print(f"--- Sheet: {sheet_name} ---")
        sheet = wb[sheet_name]
        for r in range(1, min(sheet.max_row+1, 30)):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, min(sheet.max_column+1, 15))]
            print(f"Row {r}: {row_vals}")

if __name__ == "__main__":
    dump_rtest()
