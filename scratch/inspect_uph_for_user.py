import openpyxl

def inspect_uph_for_user():
    s_path = r"C:\Users\jackn\spyder\amphenolrf\docs\Stest.xlsx"
    wb = openpyxl.load_workbook(s_path, data_only=True)
    sheet = wb.worksheets[0]
    
    # Let's search for item code containing "105356" or "121231"
    for r in range(3, sheet.max_row+1):
        item = str(sheet.cell(row=r, column=2).value or "").strip()
        uph = sheet.cell(row=r, column=20).value
        mo = sheet.cell(row=r, column=9).value
        qty = sheet.cell(row=r, column=10).value
        if "105356" in item or "121231" in item:
            print(f"Row {r} | Item: {item} | MO: {mo} | Qty: {qty} | UPH: {uph}")

if __name__ == "__main__":
    inspect_uph_for_user()
