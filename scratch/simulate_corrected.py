import openpyxl

def simulate_corrected():
    s_path = r"C:\Users\jackn\spyder\amphenolrf\docs\Stest.xlsx"
    r_path = r"C:\Users\jackn\spyder\amphenolrf\docs\Rtest.xlsx"

    # Load resources
    r_wb = openpyxl.load_workbook(r_path, data_only=True)
    r_sheet = r_wb.worksheets[0]
    line_capacities = {}
    
    for row in r_sheet.iter_rows(values_only=True):
        if not row or len(row) == 0:
            continue
        cell_0 = str(row[0] or "").strip().upper()
        if cell_0.startswith("L") and cell_0[1:].isdigit():
            capacities = []
            for c in range(1, 8):
                val = 0.0
                if c < len(row) and row[c] is not None:
                    try:
                        val = float(row[c])
                    except:
                        pass
                capacities.append(val)
            line_capacities[cell_0] = capacities

    active_lines = ["L1", "L2", "L3", "L4", "L5", "L6", "L7", "L8", "L9", "L10", "L11", "L12", "L14", "L15", "L16"]
    for line in active_lines:
        if line not in line_capacities:
            line_capacities[line] = [10.5, 10.5, 10.5, 10.5, 10.5, 10.5, 10.5]

    s_wb = openpyxl.load_workbook(s_path, data_only=True)
    sheet = s_wb.worksheets[0]
    
    rows = list(sheet.iter_rows(values_only=True))
    data_rows = rows[2:]
    
    parsed_items = []
    
    def text_val(val):
        if val is None:
            return ""
        return str(val).strip()

    def num_val(val, fallback=0.0):
        if val is None:
            return fallback
        try:
            return float(val)
        except:
            return fallback

    def normalize(val):
        import unicodedata
        text = text_val(val).lower()
        text = unicodedata.normalize('NFD', text)
        text = "".join([c for c in text if unicodedata.category(c) != 'Mn'])
        text = text.replace('đ', 'd')
        return text

    def is_not_no(val):
        return val != "no"

    for r_idx, row in enumerate(data_rows):
        row_number = r_idx + 3
        if not row or len(row) < 10:
            continue
        
        # CORRECTED mappings:
        # B (index 1) is Item
        # C (index 2) is Item description
        item = text_val(row[1])
        desc = text_val(row[2])
        mo = text_val(row[8])
        quantity = num_val(row[9])
        uph = num_val(row[19])
        status = text_val(row[7])
        pmc = text_val(row[6])
        
        if not quantity and not mo and not item and not desc:
            continue
        
        if uph <= 0:
            uph = 300.0

        solder_pn = text_val(row[58]) # Corrected (Index 58 is BG)
        ink_printing = text_val(row[60]) # Corrected (Index 60 is BI)
        letter_rolling = text_val(row[61]) # Corrected (Index 61 is BJ)
        tape_reel = text_val(row[62]) # Corrected (Index 62 is BK)
        
        norm_desc = normalize(desc)
        norm_status = normalize(status)
        norm_pmc = normalize(pmc)
        norm_solder = normalize(solder_pn)
        norm_ink = normalize(ink_printing)
        norm_letter = normalize(letter_rolling)
        norm_tape = normalize(tape_reel)
        
        pmc_score = 6 if "uu tien 1" in norm_pmc else 1
        sap_score = 5 if "sap" in norm_desc or is_not_no(norm_letter) else 1
        kap_score = 5 if "kap" in norm_desc or is_not_no(norm_ink) else 1
        
        is_lathe = item.upper().startswith("34-")
        lathe_score = 5 if is_lathe else 1
        
        reel_score = 4 if "smp" in norm_desc or "stpk" in norm_desc or is_not_no(norm_tape) else 1
        solder_score = 3 if norm_solder == "yes" else 1
        large_qty_score = 2 if quantity > 5000 else 1
        
        material_rank = 1
        try:
            h_val = int(status.strip())
            if h_val == 3: material_rank = 4
            elif h_val == 2: material_rank = 3
            elif h_val == 1: material_rank = 2
            elif h_val == 0: material_rank = 1
        except:
            if "ready material" in norm_status:
                material_rank = 4
            elif "co lieu" in norm_status:
                material_rank = 3
                
        priority_score = pmc_score * sap_score * kap_score * lathe_score * reel_score * solder_score * large_qty_score
        
        promise_date = row[29] if len(row) > 29 else ""
        ship_date = row[25] if len(row) > 25 else ""
        
        def date_val(val):
            if isinstance(val, (int, float)):
                return float(val)
            return 999999.0
            
        parsed_items.append({
            'row_num': row_number,
            'mo': mo,
            'item': item,
            'desc': desc,
            'quantity': quantity,
            'uph': uph,
            'priority_score': priority_score,
            'material_rank': material_rank,
            'pmc_score': pmc_score,
            'promise_date': date_val(promise_date),
            'ship_date': date_val(ship_date),
            'original_order': len(parsed_items)
        })

    schedulable = [it for it in parsed_items if it['quantity'] > 0 and it['uph'] > 0]
    
    # Group by item code
    groups_map = {}
    for item_data in schedulable:
        group_key = item_data['item']
        if group_key not in groups_map:
            groups_map[group_key] = {
                'key': group_key,
                'items': [],
                'best_item': item_data,
                'first_order': item_data['original_order']
            }
        groups_map[group_key]['items'].append(item_data)
        groups_map[group_key]['first_order'] = min(groups_map[group_key]['first_order'], item_data['original_order'])
        
        def is_better(a, b):
            if a['priority_score'] != b['priority_score']:
                return a['priority_score'] > b['priority_score']
            if a['material_rank'] != b['material_rank']:
                return a['material_rank'] > b['material_rank']
            if a['pmc_score'] != b['pmc_score']:
                return a['pmc_score'] > b['pmc_score']
            if a['promise_date'] != b['promise_date']:
                return a['promise_date'] < b['promise_date']
            if a['ship_date'] != b['ship_date']:
                return a['ship_date'] < b['ship_date']
            return a['original_order'] < b['original_order']
            
        if is_better(item_data, groups_map[group_key]['best_item']):
            groups_map[group_key]['best_item'] = item_data

    plan_groups = list(groups_map.values())
    
    def group_key_fn(g):
        best = g['best_item']
        return (
            -best['priority_score'],
            -best['material_rank'],
            -best['pmc_score'],
            best['promise_date'],
            best['ship_date'],
            best['original_order']
        )
        
    plan_groups.sort(key=group_key_fn)

    print(f"[CORRECTED] Number of product groups (unique Items): {len(plan_groups)}")

    allocated_times = {line: [0.0]*7 for line in active_lines}
    line_pointer = 0
    group_assignments = {}
    
    for g in plan_groups:
        assigned_line = active_lines[line_pointer]
        group_assignments[g['key']] = assigned_line
        line_pointer = (line_pointer + 1) % len(active_lines)
        
        capacities = line_capacities[assigned_line]
        preferred_day = 0
        for d in range(7):
            if capacities[d] - allocated_times[assigned_line][d] > 0.01:
                preferred_day = d
                break
                
        g['items'].sort(key=lambda x: (
            -x['priority_score'],
            -x['material_rank'],
            -x['pmc_score'],
            x['promise_date'],
            x['ship_date'],
            x['original_order']
        ))
        
        days_to_allocate = list(range(preferred_day, 7))
        qty_efficiency = 1.0
        lot_step = 500.0
        
        for item in g['items']:
            remaining = item['quantity']
            uph_eff = item['uph'] * qty_efficiency
            
            allocated_qty_days = [0.0]*7
            
            for day_idx in days_to_allocate:
                if remaining <= 0:
                    break
                cap = capacities[day_idx]
                used = allocated_times[assigned_line][day_idx]
                avail = max(cap - used, 0.0)
                
                if avail <= 0.001:
                    continue
                    
                max_qty = avail * uph_eff
                chunk = min(remaining, max_qty)
                
                if lot_step > 1 and remaining > lot_step and chunk >= lot_step:
                    rounded = max(lot_step, round(chunk / lot_step) * lot_step)
                    chunk = min(rounded, remaining, max_qty)
                else:
                    chunk = min(chunk, remaining)
                    
                if chunk > 0:
                    time_used = chunk / uph_eff
                    allocated_qty_days[day_idx] += chunk
                    allocated_times[assigned_line][day_idx] += time_used
                    remaining -= chunk
                    
            item['allocated_qty'] = sum(allocated_qty_days)
            item['missing'] = remaining
            item['assigned_line'] = assigned_line

    print("\n--- [CORRECTED] LINE CAPACITY UTILIZATION SUMMARY ---")
    line_groups_count = {line: 0 for line in active_lines}
    line_total_qty_demanded = {line: 0.0 for line in active_lines}
    line_total_qty_scheduled = {line: 0.0 for line in active_lines}
    
    for g in plan_groups:
        line = group_assignments[g['key']]
        line_groups_count[line] += 1
        for item in g['items']:
            line_total_qty_demanded[line] += item['quantity']
            line_total_qty_scheduled[line] += item['allocated_qty']

    for line in active_lines:
        cap_sum = sum(line_capacities[line])
        used_sum = sum(allocated_times[line])
        demanded = line_total_qty_demanded[line]
        scheduled = line_total_qty_scheduled[line]
        missing = demanded - scheduled
        print(f"Line {line:3s}: Groups={line_groups_count[line]:2d} | Cap={cap_sum:5.1f}h | Used={used_sum:5.1f}h | Demanded={demanded:8.1f} | Scheduled={scheduled:8.1f} | Missing={missing:8.1f}")

if __name__ == "__main__":
    simulate_corrected()
