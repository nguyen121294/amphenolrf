import openpyxl

def unique_vals():
    s_path = r"C:\Users\jackn\spyder\amphenolrf\docs\Stest.xlsx"
    wb = openpyxl.load_workbook(s_path, data_only=True)
    sheet = wb.worksheets[0]
    
    col_d = set()
    col_h = set()
    col_ak = set()
    col_an = set()
    
    for r in range(3, sheet.max_row+1):
        col_d.add(sheet.cell(row=r, column=4).value)
        col_h.add(sheet.cell(row=r, column=8).value)
        col_ak.add(sheet.cell(row=r, column=37).value)
        col_an.add(sheet.cell(row=r, column=40).value)
        
    print("Unique D:", col_d)
    print("Unique H:", col_h)
    print("Unique AK:", col_ak)
    print("Unique AN:", col_an)

if __name__ == "__main__":
    unique_vals()
