import openpyxl

def dump_detailed():
    s_path = r"C:\Users\jackn\spyder\amphenolrf\docs\Stest.xlsx"
    r_path = r"C:\Users\jackn\spyder\amphenolrf\docs\Rtest.xlsx"

    # Load resources
    r_wb = openpyxl.load_workbook(r_path, data_only=True)
    r_sheet = r_wb.worksheets[0]
    line_capacities = {}
    
    for row in r_sheet.iter_rows(values_only=True):
        if not row or len(row) == 0:
            continue
        cell_0 = str(row[0] or "").strip().upper()
        if cell_0.startswith("L") and cell_0[1:].isdigit():
            capacities = []
            for c in range(1, 8):
                val = 0.0
                if c < len(row) and row[c] is not None:
                    try:
                        val = float(row[c])
                    except:
                        pass
                capacities.append(val)
            line_capacities[cell_0] = capacities

    # Parse active_lines dynamically from Excel
    parsed_lines = list(line_capacities.keys())
    import re
    def natural_sort_key(s):
        return [int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', s)]
    active_lines = sorted(parsed_lines, key=natural_sort_key)

    s_wb = openpyxl.load_workbook(s_path, data_only=True)
    sheet = s_wb.worksheets[0]
    print(f"Sheet: {sheet.title}")
    
    # Let's inspect the first row containing column headers to check if we parsed the right columns!
    # Usually header is at row 2 (index 1)
    headers = [str(cell).strip() if cell else "" for cell in next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))]
    print("Row 2 headers:")
    for idx, h in enumerate(headers):
        if h:
            print(f"Col {idx} ({openpyxl.utils.get_column_letter(idx+1)}): {h}")
            
if __name__ == "__main__":
    dump_detailed()
