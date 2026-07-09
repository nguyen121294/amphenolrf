import openpyxl

def search_terms():
    s_path = r"C:\Users\jackn\spyder\amphenolrf\docs\Stest.xlsx"
    wb = openpyxl.load_workbook(s_path, data_only=True)
    sheet = wb.worksheets[0]
    
    found = False
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        for c_idx, val in enumerate(row):
            if val is not None:
                val_str = str(val).lower()
                if "co lieu" in val_str or "ready" in val_str or "material" in val_str:
                    col_letter = openpyxl.utils.get_column_letter(c_idx + 1)
                    print(f"Row {r_idx+1}, Col {col_letter} ({c_idx}): {val}")
                    found = True
    if not found:
        print("No matching terms found.")

if __name__ == "__main__":
    search_terms()
