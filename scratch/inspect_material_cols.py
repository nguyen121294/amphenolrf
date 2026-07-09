import openpyxl

def inspect_material_cols():
    s_path = r"C:\Users\jackn\spyder\amphenolrf\docs\Stest.xlsx"
    wb = openpyxl.load_workbook(s_path, data_only=True)
    sheet = wb.worksheets[0]
    
    print("Row 2, V and W:")
    print("V2:", sheet.cell(row=2, column=22).value)
    print("W2:", sheet.cell(row=2, column=23).value)
    
    # Let's see Row 3 to 15
    for r in range(3, 16):
        v_val = sheet.cell(row=r, column=22).value
        w_val = sheet.cell(row=r, column=23).value
        print(f"Row {r} | V (21): {v_val} | W (22): {w_val}")

if __name__ == "__main__":
    inspect_material_cols()
